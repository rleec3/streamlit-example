import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
from lxml import etree
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title='Nonprofit Search Tool', page_icon='C3_Only_Ball.png', layout='wide')

# Function to fetch years and corresponding URLs for the given EIN
def fetch_years(ein):
    base_url = "https://projects.propublica.org"
    url = f"{base_url}/nonprofits/organizations/{ein}"
    response = requests.get(url)
    years = {}
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        sections = soup.find_all("section", class_="single-filing-period")
        for section in sections:
            year = section['id'].replace("filing", "")
            links = section.find_all("a", class_="btn")
            xml_link = None
            for link in links:
                if 'xml' in link.text.lower():
                    xml_link = link['href']
                    break
            if xml_link:
                years[year] = base_url + xml_link
                object_id = xml_link.split('object_id=')[-1]
                detailed_url = f"{base_url}/nonprofits/download-xml?object_id={object_id}"
                years[year] = (xml_link, detailed_url)
            else:
                years[year] = "XML link not found"
    return years

# Define namespaces for XML parsing
ns = {'efile': 'http://www.irs.gov/efile'}

def get_text(element, xpaths, namespaces):
    for xpath in xpaths:
        result = element.xpath(xpath, namespaces=namespaces)
        if result:
            return result[0]
    return "Not Available"

def fetch_data(ein, detailed_url):
    response = requests.get(detailed_url)
    if response.status_code == 200:
        tree = etree.fromstring(response.content)
        
        organization_data = {
            'EIN': ein,
            'Business Name': get_text(tree, ['//efile:Return/efile:ReturnHeader/efile:Filer/efile:BusinessName/efile:BusinessNameLine1Txt/text()'], ns),
            'City': get_text(tree, ['//efile:Return/efile:ReturnHeader/efile:Filer/efile:USAddress/efile:CityNm/text()'], ns),
            'State': get_text(tree, ['//efile:Return/efile:ReturnHeader/efile:Filer/efile:USAddress/efile:StateAbbreviationCd/text()'], ns),
            'Fiscal Year End': get_text(tree, ['//efile:Return/efile:ReturnHeader/efile:TaxPeriodEndDt/text()'], ns),
            'Total Assets EOY': get_text(tree, ['//efile:Return/efile:ReturnData/efile:IRS990/efile:TotalAssetsEOYAmt/text()', './/efile:FMVAssetsEOYAmt/text()'], ns),
            'Total Expenses': get_text(tree, ['//efile:Return/efile:ReturnData/efile:IRS990/efile:CYTotalExpensesAmt/text()', './/efile:TotalExpensesRevAndExpnssAmt/text()'], ns),
            'Total Revenue': get_text(tree, ['//efile:Return/efile:ReturnData/efile:IRS990/efile:CYTotalRevenueAmt/text()', './/efile:TotalRevAndExpnssAmt/text()'], ns),
            'Employee Count': get_text(tree, ['//efile:Return/efile:ReturnData/efile:IRS990/efile:TotalEmployeeCnt/text()'], ns)
        }

        fiscal_year_end_text = organization_data["Fiscal Year End"]
        if "-" in fiscal_year_end_text:
            fiscal_year_parts = fiscal_year_end_text.split("-")
            if len(fiscal_year_parts) == 3:
                fiscal_year_month = int(fiscal_year_parts[1])
                fiscal_year_year = int(fiscal_year_parts[0])
                if fiscal_year_month == 12:
                    w_year_end = fiscal_year_end_text
                else:
                    w_year_end = f"{fiscal_year_year - 1}-12-31"
                organization_data["WYearEnd"] = w_year_end

        individuals_data = []
        part_j_sections = tree.xpath('//efile:Return/efile:ReturnData/efile:IRS990ScheduleJ/efile:RltdOrgOfficerTrstKeyEmplGrp', namespaces=ns)
        for section in part_j_sections:
            name = get_text(section, ['.//efile:PersonNm/text()', './/efile:BusinessNameLine1Txt/text()'], ns)
            title = section.xpath('.//efile:TitleTxt/text()', namespaces=ns)[0] if section.xpath('.//efile:TitleTxt/text()', namespaces=ns) else "Not Available"
            compensation_data = {
                'Base Compensation': section.xpath('.//efile:BaseCompensationFilingOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:BaseCompensationFilingOrgAmt/text()', namespaces=ns) else "Not Available",
                'Bonus': section.xpath('.//efile:BonusFilingOrganizationAmount/text()', namespaces=ns)[0] if section.xpath('.//efile:BonusFilingOrganizationAmount/text()', namespaces=ns) else "Not Available",
                'Other Compensation': section.xpath('.//efile:OtherCompensationFilingOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:OtherCompensationFilingOrgAmt/text()', namespaces=ns) else "Not Available",
                'Deferred Compensation': section.xpath('.//efile:DeferredCompensationFlngOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:DeferredCompensationFlngOrgAmt/text()', namespaces=ns) else "Not Available",
                'Nontaxable Benefits': section.xpath('.//efile:NontaxableBenefitsFilingOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:NontaxableBenefitsFilingOrgAmt/text()', namespaces=ns) else "Not Available",
                'Total Compensation': section.xpath('.//efile:TotalCompensationFilingOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:TotalCompensationFilingOrgAmt/text()', namespaces=ns) else "Not Available",
                'Reportable Compensation (Part VII)': section.xpath('.//efile:ReportableCompFromOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:ReportableCompFromOrgAmt/text()', namespaces=ns) else "Not Available",
                'Reportable Compensation From Rltd Org (Part VII)': section.xpath('.//efile:ReportableCompFromRltdOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:ReportableCompFromRltdOrgAmt/text()', namespaces=ns) else "Not Available",
                'Other Compensation (Part VII)': section.xpath('.//efile:OtherCompensationAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:OtherCompensationAmt/text()', namespaces=ns) else "Not Available",
                'Avg Hr Per Week (Part VII) Test': section.xpath('.//efile:AverageHoursPerWeekRt/text()', namespaces=ns)[0] if section.xpath('.//efile:AverageHoursPerWeekRt/text()', namespaces=ns) else "Not Available",
            }
            individual_data = {'Name': name, 'Title': title}
            individual_data.update(compensation_data)
            individuals_data.append(individual_data)
        
        individuals_data2 = []
        part_vii_sections = tree.xpath('//efile:Return/efile:ReturnData/efile:IRS990/efile:Form990PartVIISectionAGrp', namespaces=ns)
        for section in part_vii_sections:
            name = get_text(section, ['.//efile:PersonNm/text()', './/efile:BusinessNameLine1Txt/text()'], ns)
            title = get_text(section, ['.//efile:TitleTxt/text()', './/efile:TitleTxt/text()'], ns)
            compensation_data2 = {
                'Reportable Compensation (Part VII)': section.xpath('.//efile:ReportableCompFromOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:ReportableCompFromOrgAmt/text()', namespaces=ns) else "Not Available",
                'Reportable Compensation From Rltd Org (Part VII)': section.xpath('.//efile:ReportableCompFromRltdOrgAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:ReportableCompFromRltdOrgAmt/text()', namespaces=ns) else "Not Available",
                'Other Compensation (Part VII)': section.xpath('.//efile:OtherCompensationAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:OtherCompensationAmt/text()', namespaces=ns) else "Not Available",
                'Avg Hr Per Week (Part VII)': section.xpath('.//efile:AverageHoursPerWeekRt/text()', namespaces=ns)[0] if section.xpath('.//efile:AverageHoursPerWeekRt/text()', namespaces=ns) else "Not Available",
            }
            individual_data2 = {'Name': name, 'Title': title}
            individual_data2.update(compensation_data2)
            individuals_data2.append(individual_data2)
        
        individuals_data3 = []
        part_pf_sections = tree.xpath('//efile:Return/efile:ReturnData/efile:IRS990PF/efile:OfficerDirTrstKeyEmplInfoGrp/efile:OfficerDirTrstKeyEmplGrp', namespaces=ns)
        for section in part_pf_sections:
            name = get_text(section, ['.//efile:PersonNm/text()', './/efile:BusinessNameLine1Txt/text()'], ns)
            title = get_text(section, ['.//efile:TitleTxt/text()'], ns)
            compensation_data3 = {
                'Reportable Compensation (PF)': section.xpath('.//efile:CompensationAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:CompensationAmt/text()', namespaces=ns) else "Not Available",
                'Employee Benefit Amount (PF)': section.xpath('.//efile:EmployeeBenefitProgramAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:EmployeeBenefitProgramAmt/text()', namespaces=ns) else "Not Available",
                'Other Compensation (PF)': section.xpath('.//efile:ExpenseAccountOtherAllwncAmt/text()', namespaces=ns)[0] if section.xpath('.//efile:ExpenseAccountOtherAllwncAmt/text()', namespaces=ns) else "Not Available",
                'Avg Hr Per Week (PF)': section.xpath('.//efile:AverageHrsPerWkDevotedToPosRt/text()', namespaces=ns)[0] if section.xpath('.//efile:AverageHrsPerWkDevotedToPosRt/text()', namespaces=ns) else "Not Available",
            }
            individual_data3 = {'Name': name, 'Title': title}
            individual_data3.update(compensation_data3)
            individuals_data3.append(individual_data3)
        
        combined_individuals_data = individuals_data + individuals_data2 + individuals_data3
        unique_individuals = {individual['Name']: individual for individual in combined_individuals_data}.values()

        final_individuals_data = []
        for individual in unique_individuals:
            merged_data = {}
            for dataset in [individuals_data, individuals_data2, individuals_data3]:
                for data in dataset:
                    if data['Name'] == individual['Name']:
                        merged_data.update(data)
            final_individuals_data.append(merged_data)

        return {'organization_data': organization_data, 'individuals_data': final_individuals_data}

def edit_excel_template(data, template_path, num_orgs, num_years):
    def to_number(value):
        try:
            return float(value)
        except ValueError:
            return value

    def to_proper_case(text):
        return text.title()

    def to_date(value, date_format="%Y-%m-%d"):
        try:
            return datetime.strptime(value, date_format)
        except ValueError:
            return value

    def copy_row_formatting_and_values(sheet, source_row, target_row):
        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            source_cell = sheet[f"{col_letter}{source_row}"]
            target_cell = sheet[f"{col_letter}{target_row}"]
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.value = source_cell.value

    def insert_blank_row_with_formatting(sheet, row):
        sheet.insert_rows(row)
        copy_row_formatting_and_values(sheet, row + 1, row)

    def is_merged_cell(sheet, row, col):
        cell = f"{get_column_letter(col)}{row}"
        for merged_range in sheet.merged_cells.ranges:
            if cell in merged_range:
                return True
        return False

    def get_top_left_merged_cell(sheet, row, col):
        cell = f"{get_column_letter(col)}{row}"
        for merged_range in sheet.merged_cells.ranges:
            if cell in merged_range:
                return merged_range.min_row, merged_range.min_col
        return row, col

    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook["Form 990 - Position Title"]
    sheet2 = workbook["PEER GROUP"]
    sheet4 = workbook["Form 990PF - Position Title"]
    sheet_setup = workbook["SETUP"]
    start_row_peer_group = 6  # The row to start inserting in PEER GROUP
    start_row_990 = 6  # The row to start inserting in Form 990 - Position Title
    start_row_990pf = 6  # The row to start inserting in Form 990PF - Position Title
    index_counter = 1  # Index counter for PEER GROUP

    unique_ein_years = set()
    entries_to_insert = []

    # Collect unique EIN/year entries
    for entry in data:
        unique_key = (entry["EIN"], entry["W2E"])
        if unique_key not in unique_ein_years:
            unique_ein_years.add(unique_key)
            entries_to_insert.append(entry)

    # Insert the required number of rows into the PEER GROUP tab
    for _ in range(len(entries_to_insert)):
        insert_blank_row_with_formatting(sheet2, start_row_peer_group)
        start_row_peer_group += 1

    # Fill in the PEER GROUP tab with data
    start_row_peer_group = 6  # Reset to the starting row for data insertion
    for entry in entries_to_insert:
        sheet2[f"A{start_row_peer_group}"] = index_counter
        sheet2[f"B{start_row_peer_group}"] = to_proper_case(entry["Organization_Name"])
        
        # Handle merged cells
        if not is_merged_cell(sheet2, start_row_peer_group, 3):
            sheet2[f"C{start_row_peer_group}"] = to_number(entry["EIN"])
        else:
            top_left_row, top_left_col = get_top_left_merged_cell(sheet2, start_row_peer_group, 3)
            if top_left_row == start_row_peer_group and top_left_col == 3:
                sheet2[f"C{start_row_peer_group}"] = to_number(entry["EIN"])

        sheet2[f"F{start_row_peer_group}"] = to_proper_case(entry["City"])
        sheet2[f"G{start_row_peer_group}"] = entry["State"]
        sheet2[f"E{start_row_peer_group}"] = to_date(entry.get("W2E", entry.get("WYearEnd", "Not Available")))
        sheet2[f"D{start_row_peer_group}"] = to_date(entry.get("Fiscal_Year_End", "Not Available"))
        sheet2[f"H{start_row_peer_group}"] = to_number(entry.get("Total Assets", "Not Available"))
        sheet2[f"I{start_row_peer_group}"] = to_number(entry.get("Total Expenses", "Not Available"))
        sheet2[f"J{start_row_peer_group}"] = to_number(entry.get("Total Revenue", "Not Available"))
        sheet2[f"N{start_row_peer_group}"] = to_number(entry.get("Employee Count", "Not Available"))
        sheet2[f"K{start_row_peer_group}"] = to_number(entry["Total Assets"]) / 1000000
        sheet2[f"L{start_row_peer_group}"] = to_number(entry["Total Expenses"]) / 1000000
        sheet2[f"M{start_row_peer_group}"] = to_number(entry["Total Revenue"]) / 1000000
        start_row_peer_group += 1
        index_counter += 1

    # Generate entries for Form 990 and Form 990PF with individual data
    index_counter = 1
    for entry in data:
        #for individual in entry.get('individuals_data', []):
        insert_blank_row_with_formatting(sheet, start_row_990)

        sheet[f"C{start_row_990}"] = to_number(entry["EIN"])
        sheet[f"B{start_row_990}"] = to_proper_case(entry["Organization_Name"])
        sheet[f"F{start_row_990}"] = to_proper_case(entry["City"])
        sheet[f"G{start_row_990}"] = entry["State"]
        sheet[f"E{start_row_990}"] = to_date(entry.get("W2E", entry.get("WYearEnd", "Not Available")))
        sheet[f"D{start_row_990}"] = to_date(entry.get("Fiscal_Year_End", "Not Available"))
        sheet[f"J{start_row_990}"] = to_number(entry["Total Assets"]) / 1000000
        sheet[f"K{start_row_990}"] = to_number(entry["Total Expenses"]) / 1000000
        sheet[f"L{start_row_990}"] = to_number(entry["Total Revenue"]) / 1000000
        sheet[f"M{start_row_990}"] = to_number(entry.get("Employee Count", "Not Available"))

        sheet[f"H{start_row_990}"] = to_proper_case(entry["Employee_Name"])
        sheet[f"I{start_row_990}"] = to_proper_case(entry["Title_Of_Position"])
        sheet[f"P{start_row_990}"] = to_number(entry["Base Compensation"]) / 1000
        sheet[f"T{start_row_990}"] = to_number(entry["Deferred Compensation"]) / 1000
        sheet[f"S{start_row_990}"] = to_number(entry["Other Compensation"]) / 1000
        sheet[f"U{start_row_990}"] = to_number(entry["Nontaxable Benefits"]) / 1000
        sheet[f"Q{start_row_990}"] = to_number(entry["Bonus"]) / 1000
        sheet[f"R{start_row_990}"] = (to_number(entry["Bonus"]) / 1000) + (to_number(entry["Base Compensation"]) / 1000)
        sheet[f"V{start_row_990}"] = (to_number(entry["Bonus"]) / 1000) + (to_number(entry["Base Compensation"]) / 1000) + (to_number(entry["Other Compensation"]) / 1000) + (to_number(entry["Deferred Compensation"]) / 1000) + to_number(entry["Nontaxable Benefits"]) / 1000

        insert_blank_row_with_formatting(sheet4, start_row_990pf)

        sheet4[f"C{start_row_990pf}"] = to_number(entry["EIN"])
        sheet4[f"B{start_row_990pf}"] = to_proper_case(entry["Organization_Name"])
        sheet4[f"F{start_row_990pf}"] = to_proper_case(entry["City"])
        sheet4[f"G{start_row_990pf}"] = entry["State"]
        sheet4[f"E{start_row_990pf}"] = to_date(entry.get("W2E", entry.get("WYearEnd", "Not Available")))
        sheet4[f"D{start_row_990pf}"] = to_date(entry.get("Fiscal_Year_End", "Not Available"))
        sheet4[f"J{start_row_990pf}"] = to_number(entry["Total Assets"]) / 1000000
        sheet4[f"K{start_row_990pf}"] = to_number(entry["Total Expenses"]) / 1000000
        
        sheet4[f"L{start_row_990pf}"] = to_number(entry.get("Employee Count", "Not Available"))

        sheet4[f"H{start_row_990pf}"] = to_proper_case(entry["Employee_Name"])
        sheet4[f"I{start_row_990pf}"] = to_proper_case(entry["Title_Of_Position"])
        sheet4[f"O{start_row_990pf}"] = to_number(entry["Reportable Comp PF"]) / 1000
        sheet4[f"P{start_row_990pf}"] = to_number(entry["Benefits Comp PF"]) / 1000
        sheet4[f"Q{start_row_990pf}"] = to_number(entry["Expenses and Other Comp PF"]) / 1000
        sheet4[f"R{start_row_990pf}"] = to_number(entry["Expenses and Other Comp PF"]) / 1000 +  (to_number(entry["Benefits Comp PF"]) / 1000) + (to_number(entry["Reportable Comp PF"]) / 1000)

        start_row_990 += 1
        start_row_990pf += 1

        index_counter += 1

    sheet_setup["C4"] = to_proper_case(target_org_data["Business Name"])
    sheet_setup["C9"] = to_proper_case(target_org_data["City"])
    sheet_setup["C10"] = target_org_data["State"]
    sheet_setup["C11"] = to_number(target_org_data["Total Assets EOY"]) 
    sheet_setup["C12"] = to_number(target_org_data["Total Revenue"]) 
    sheet_setup["C13"] = to_number(target_org_data["Total Expenses"]) 
    sheet_setup["C14"] = to_number(target_org_data["Employee Count"])
    # Set row height for all sheets except SETUP
    for sheet_name in workbook.sheetnames:
        if sheet_name != "SETUP":
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                sheet.row_dimensions[row[0].row].height = 30

    edited_file = BytesIO()
    workbook.save(edited_file)
    edited_file.seek(0)
    return edited_file

# Additional helper functions
def is_merged_cell(sheet, row, col):
    cell = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell in merged_range:
            return True
    return False

def get_top_left_merged_cell(sheet, row, col):
    cell = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell in merged_range:
            return merged_range.min_row, merged_range.min_col
    return row, col

# Streamlit UI components
banner_path = 'Horizontal_Banner_NoSC.png'
st.image(banner_path, width=400)
st.header("C3 990 Tool Edit/Upload")
st.write("Instructions: Copy and paste a list of EINs below (one per line) and select the desired number of years to fetch data for each organization.")

# User inputs EINs directly
eins_input = st.text_area("Paste EINs here (one per line):", height=200)
target_ein_input = st.text_input("Enter Target Org EIN:")
if eins_input or target_ein_input:
    eins = [ein.strip() for ein in eins_input.split('\n') if ein.strip()]
    num_orgs = len(eins)
    st.write(f"Found {num_orgs} EINs in the input.")
    
    num_years = st.number_input("Select the number of years to fetch data for", min_value=1, max_value=5, value=1, step=1)
    individuals_option = st.selectbox("Select individuals to pull", ["Top 5 Paid Individuals", "All Individuals"])

    st.session_state['organizations_data'] = []
    st.session_state['all_individuals_data'] = []
    st.session_state['year_data'] = {}
    st.session_state['selected_years'] = {}

    for i in range(num_orgs):
        st.session_state['year_data'][str(i)] = fetch_years(eins[i])

    target_org_data = None
    if target_ein_input:
        target_year_data = fetch_years(target_ein_input)
        if target_year_data:
            target_years = sorted(target_year_data.keys(), reverse=True)[:num_years]
            detailed_url = target_year_data[target_years[0]][1]
            target_org_data = fetch_data(target_ein_input, detailed_url)['organization_data']

    if st.button("Generate Final Output Chart"):
        st.session_state['organizations_data'] = []
        st.session_state['all_individuals_data'] = []
        progress_text = st.empty()
        progress_bar = st.progress(0)
        for i, ein in enumerate(eins):
            try:
                if ein.strip() and st.session_state['year_data'][str(i)]:
                    years = list(st.session_state['year_data'][str(i)].keys())
                    selected_years = sorted(years, reverse=True)[:num_years]
                    for year in selected_years:
                        st.session_state['selected_years'][str(i)] = year
                        detailed_url = st.session_state['year_data'][str(i)][year][1]
                        fetched_data = fetch_data(ein, detailed_url)
                        organization_data = fetched_data['organization_data']
                        organization_data['individuals_data'] = fetched_data['individuals_data']
                        st.session_state['organizations_data'].append(organization_data)
            except Exception as e:
                st.write(f"Error fetching data for EIN {ein}: {e}")
                st.session_state['organizations_data'].append({"EIN": ein, "Business Name": "Not Found", 'individuals_data': []})
            progress = (i + 1) / num_orgs
            progress_bar.progress(progress)
            progress_text.text(f"{i + 1}/{num_orgs} EINs Parsed")

        final_chart_data = []
        st.session_state['final_chart_data'] = []
        existing_data_identifiers = {(row['EIN'], row['Employee_Name']) for row in st.session_state['final_chart_data']}
        for organization_data in st.session_state['organizations_data']:
            individuals_data = organization_data['individuals_data']

            for individual_data in individuals_data:
                try:
                    individual_data['Total Compensation (PF)'] = (
                        float(individual_data.get('Reportable Compensation (PF)', 0)) +
                        float(individual_data.get('Employee Benefit Amount (PF)', 0)) +
                        float(individual_data.get('Other Compensation (PF)', 0))
                    )
                except Exception as e:
                    individual_data['Total Compensation (PF)'] = 0
                    st.write(f"Error calculating Total Compensation (PF) for {individual_data.get('Name')}: {e}")
                try:
                    individual_data['Total Compensation'] = float(individual_data.get('Total Compensation', 0))
                except Exception as e:
                    individual_data['Total Compensation'] = 0
                    st.write(f"Error calculating Total Compensation for {individual_data.get('Name')}: {e}")

            if individuals_option == "Top 5 Paid Individuals":
                selected_individuals = sorted(individuals_data, key=lambda x: x['Total Compensation (PF)'] + x['Total Compensation'], reverse=True)[:5]
            else:
                selected_individuals = individuals_data

            for individual_data in selected_individuals:
                name = individual_data['Name']
                title = individual_data['Title']
                if (organization_data.get('EIN'), name) not in existing_data_identifiers:
                    chart_row = {
                        "Organization_Name": organization_data.get('Business Name', 'Unknown'),
                        "EIN": organization_data.get('EIN', 'Not Available'),
                        "W2E": organization_data.get('WYearEnd', 'Not Available'),
                        "Fiscal_Year_End": organization_data.get('Fiscal Year End', 'Not Available'),
                        "City": organization_data.get('City', 'Not Available'),
                        "State": organization_data.get('State', 'Not Available'),
                        "Employee_Name": name,
                        "Title_Of_Position": title,
                        "Total Assets": organization_data.get('Total Assets EOY', 'Not Available'),
                        "Total Expenses": organization_data.get('Total Expenses', 'Not Available'),
                        "Total Revenue": organization_data.get('Total Revenue', 'Not Available'),
                        "Employee Count": organization_data.get('Employee Count', 'Not Available'),
                        "Base Compensation": individual_data.get('Base Compensation', '0'),
                        "Bonus": individual_data.get('Bonus', '0'),
                        "Other Compensation": individual_data.get('Other Compensation','0'),
                        "Deferred Compensation": individual_data.get('Deferred Compensation', '0'),
                        "Nontaxable Benefits": individual_data.get('Nontaxable Benefits', '0'),
                        "Total Compensation": individual_data.get('Total Compensation', '0'),
                        "Reportable Comp PF": individual_data.get('Reportable Compensation (PF)', '0'),
                        "Benefits Comp PF": individual_data.get('Employee Benefit Amount (PF)', '0'),
                        "Expenses and Other Comp PF": individual_data.get('Other Compensation (PF)', '0'),
                        "Total Compensation (PF)": individual_data['Total Compensation (PF)']
                    }
                    final_chart_data.append(chart_row)
                    st.session_state['final_chart_data'].append(chart_row)

        final_df = pd.DataFrame(st.session_state['final_chart_data'])
        st.write(final_df)

        # Calculate the number of rows needed for PEER GROUP tab
        num_unique_entries = len({(entry["EIN"], entry["W2E"]) for entry in st.session_state['final_chart_data']})
        if final_chart_data:
            edited_file = edit_excel_template(st.session_state['final_chart_data'], '990Template2.xlsm', num_orgs, num_years)
            st.download_button(label="Download Updated 990 Template", data=edited_file, file_name="990_template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if st.button("Reset", key='reset_button'):
    st.session_state.clear()
    st.rerun()