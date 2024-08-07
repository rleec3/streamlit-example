import altair as alt
import streamlit as st
import openpyxl
import pandas as pd
import streamlit_shadcn_ui as ui
from io import BytesIO

# Assuming 'path_to_file.xlsx' is the correct path to your Excel file
file_path = 'fbc_data_2024_V1.2.xlsx'

st.set_page_config(page_title='Living Wage Dashboard', page_icon='C3_Only_Ball.png', layout='wide')




# Function to load data
@st.cache_data
def load_data(sheet_name):
    # Using the `header` parameter to specify which row contains column names
    # `engine='openpyxl'` is specified since the default engine might not handle '.xlsm' correctly
    data = pd.read_excel(file_path, sheet_name=sheet_name, header=0, engine='openpyxl')  
    data = data.dropna(axis='columns', how='all')  # Drop columns with all NaN values
    return data

# Load the County and Metro datasets
county_data = load_data('County_Annual')  # Update with the correct sheet name for County data
metro_data = load_data('Metro_Annual')  # Update with the correct sheet name for Metro data



# Streamlit application layout
banner_path = 'Horizontal_Banner_NoSC.png'
st.image(banner_path, width=400)
#ui.badges(badge_list=[ ("Under Construction", "destructive")], class_name="flex gap-2", key="main_badges1")

st.markdown("""
<style>
.flex-center-vertical {
    display: flex;
    align-items: center; /* Aligns vertically */
    justify-content: center; /* Aligns horizontally */
    height: 100%;
}
</style>
""", unsafe_allow_html=True)



st.header("Living Wage Dashboard") 
st.caption('Source: Economic Policy Institute Family Budget Calculator, January 2024. Data are in 2023 dollars.')

col1, col2 = st.columns([2, 2])  # Here, col1 is three times wider than col2

# Place a caption in each column
with col1:
    st.caption('⚪ : Data not utilized in client deliverables')

with col2:
    st.caption('🟢 : Data utilized in client deliverables')
#ui.badges(badge_list=[ ("Under Construction", "destructive")], class_name="flex gap-2", key="main_badges1")

# Filter selection sidebar
with st.sidebar:
    data_type = st.radio("Select Data Type", ("County", "Metro"))
    selected_state = st.selectbox("Select State", county_data['State abv.'].unique())
    provider_filter = st.multiselect("Select Providers", [1, 2])
    dependents_filter = st.multiselect("Select Dependents", [0, 1, 2, 3, 4])

    # Dynamic selection for monetary columns to include in total
    monetary_columns = ['Housing', 'Food', 'Transportation', 'Healthcare', 'Other Necessities ', 'Childcare', 'Taxes']
    selected_monetary_columns = st.multiselect('Select monetary columns to include in Total:', monetary_columns)
    
# Based on data type, load the appropriate area data
if data_type == "County":
    areas = county_data[county_data['State abv.'] == selected_state]['County'].unique()
else:
    areas = metro_data[metro_data['State abv.'] == selected_state]['Areaname'].unique()

sorted_areas = sorted(areas)
selected_area = st.selectbox(f"Select {data_type} Area", sorted_areas)

#selected_area = st.selectbox(f"Select {data_type} Area", areas)

# Filter the data based on selections
if data_type == "County":
    filtered_data = county_data[
        (county_data['State abv.'] == selected_state) & 
        (county_data['County'] == selected_area) &
        (county_data['Provider'].isin(provider_filter)) &
        (county_data['Dependent'].isin(dependents_filter))  # Make sure 'Dependents' matches your data column
    ]
else:
    filtered_data = metro_data[
        (metro_data['State abv.'] == selected_state) & 
        (metro_data['Areaname'] == selected_area) &
        (metro_data['Provider'].isin(provider_filter)) &
        (metro_data['Dependent'].isin(dependents_filter))  # Make sure 'Dependents' matches your data column
    ]

# Calculate the total for selected monetary columns
if selected_monetary_columns:
    filtered_data['Total'] = filtered_data[selected_monetary_columns].astype(float).sum(axis=1)

# Exclude certain columns from the final output
columns_to_exclude = ['case_id', 'top100', 'num_counties_in_st', 'st_cost_rank', 'st_med_aff_rank', 'st_income_rank', 'top100_cost_rank', 'top100_med_faminc_rank', 'top100_med_aff_rank', 'county_fips', 'median_family_income', 'family']
columns_to_include = [col for col in filtered_data.columns if col not in columns_to_exclude]
final_output = filtered_data[columns_to_include]

final_output['Provider, Dependent Configuration'] = final_output['Provider'].astype(str) + ',' + final_output['Dependent'].astype(str)

# Display the final output in the Streamlit app
#st.header("🟢 Annualized Living Wage w/ Healthcare Credit")
st.markdown("""
    <h1>
        <span style='font-size: 20px;'>⚪</span> <!-- Emoji with larger font size -->
        <span style='font-weight: normal; font-size: 30px;'>Annualized Living Wage</span> <!-- Text with smaller font size -->
    </h1>
    """, unsafe_allow_html=True)
st.dataframe(final_output)


# Apply a 20% Healthcare credit and recalculate the total
if 'Healthcare' in final_output.columns:
    healthcare_credit_df = final_output.copy()
    healthcare_credit_df['Healthcare'] *= 0.20  # Apply the 20% credit

    # Check if 'Total' needs to be recalculated
    if 'Total' in healthcare_credit_df.columns and 'Healthcare' in selected_monetary_columns:
        healthcare_credit_df['Total'] = healthcare_credit_df[selected_monetary_columns].astype(float).sum(axis=1)

    # Display the new Living Wage table with Healthcare credit
    st.markdown("""
    <h1>
        <span style='font-size: 20px;'>🟢</span> <!-- Emoji with larger font size -->
        <span style='font-weight: normal; font-size: 30px;'>Annualized Living Wage w/ Healthcare Credit</span> <!-- Text with smaller font size -->
    </h1>
    """, unsafe_allow_html=True) 
    st.dataframe(healthcare_credit_df)

if 'Other Necessities ' in healthcare_credit_df.columns:
    thriving_wage_df = healthcare_credit_df.copy()
    thriving_wage_df['Other Necessities '] *= 2  # Double the values in 'Other Necessities'

    # Check if 'Total' needs to be recalculated
    if 'Total' in thriving_wage_df.columns and 'Other Necessities ' in selected_monetary_columns:
        # Only include the selected monetary columns for the new total
        thriving_wage_df['Total'] = thriving_wage_df[selected_monetary_columns].astype(float).sum(axis=1)

    # Display the new Thriving Wage table
    st.markdown("""
    <h1>
        <span style='font-size: 20px;'>🟢</span> <!-- Emoji with larger font size -->
        <span style='font-weight: normal; font-size: 30px;'>Annualized Thriving Wage w/ Healthcare Credit</span> <!-- Text with smaller font size -->
    </h1>
    """, unsafe_allow_html=True)  
    st.dataframe(thriving_wage_df)


# comparison chart


filtered_data['Provider, Dependent Configuration'] = filtered_data['Provider'].astype(str) + ',' + filtered_data['Dependent'].astype(str)
# Create the 'Living Wage' and 'Thriving Wage' comparison DataFrame
comparison_df = pd.DataFrame({
    'Provider, Dependent Configuration': filtered_data['Provider, Dependent Configuration'],
    'Living Wage': healthcare_credit_df['Total'] / 1000,  # Divide by 1000 to match the desired unit
    'Thriving Wage': thriving_wage_df['Total'] / 1000  # Divide by 1000 to match the desired unit
})





st.markdown(f"""
    <h1 style='font-weight: normal; font-size: 30px;'>
        Household Living Wage for {selected_area}, {selected_state}
    </h1>
    """, unsafe_allow_html=True)

st.table(comparison_df.style.format({'Living Wage': '${:,.0f}', 'Thriving Wage': '${:,.0f}'}))

# Create the bar chart comparison
long_df = comparison_df.melt('Provider, Dependent Configuration', var_name='Wage Type', value_name='Wage')

# Define color scheme
colors = ['#2078a1', '#144961']  # Light teal and dark teal colors

# Create the grouped bar chart
grouped_bar_chart = alt.Chart(long_df).mark_bar().encode(
    x=alt.X('Wage Type:N', axis=alt.Axis(title='')),
    y=alt.Y('Wage:Q', axis=alt.Axis(title='Wage (1,000s)', format=',.0f'), scale=alt.Scale(zero=True)),
    color=alt.Color('Wage Type:N', scale=alt.Scale(domain=['Living Wage', 'Thriving Wage'], range=colors)),
    tooltip=[alt.Tooltip('Wage Type:N', title='Wage Type'), alt.Tooltip('Wage:Q', title='Wage', format=',.0f')],
    order=alt.Order(
        # Sort to ensure that the bars are ordered correctly within each group
        'Wage Type:N',
        sort='ascending'
    )
).properties(
    width=alt.Step(30)  # Controls the width of the bars
).facet(
    column=alt.Column('Provider, Dependent Configuration:N', header=alt.Header(title=None))
)

# Display the chart
st.markdown(f"""
    <h1 style='font-weight: normal; font-size: 30px;'>
        Household Living Wage for {selected_area}, {selected_state}
    </h1>
    """, unsafe_allow_html=True)
st.altair_chart(grouped_bar_chart, use_container_width=True)





# Create and download modified Excel file
# Note: Replace 'some_calculation_for_living' and 'some_calculation_for_thriving' with actual calculations or source columns.



# Now, ensure these columns are included when you process and filter `final_output`
st.dataframe(comparison_df)

def save_and_load_excel():
    workbook = openpyxl.load_workbook('Living_Wage_Template.xlsx')
    sheet = workbook["LW_TW_Exhibit"]
    
    merged_cells_ranges = list(sheet.merged_cells.ranges)
    # Unmerge all merged cells temporarily
    for merged_range in merged_cells_ranges:
        sheet.unmerge_cells(str(merged_range))
    # Assuming merged cells have been handled as needed
    # Reset the DataFrame index to ensure continuity
    comparison_df.reset_index(drop=True, inplace=True)

    # Start writing from row 4 in the Excel sheet
    start_row = 4

    # Clear existing data if needed
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.value = None  # Clear the cell value

    # Write new data from the DataFrame to Excel
    for index, row in comparison_df.iterrows():
        cell_row = start_row + index  # Index is now aligned with DataFrame row
        sheet[f'A{cell_row}'].value = row['Provider, Dependent Configuration']
        sheet[f'B{cell_row}'].value = round(row['Living Wage'], 2)  # Optionally round the values
        sheet[f'C{cell_row}'].value = round(row['Thriving Wage'], 2)
    # Save the workbook to a BytesIO stream for download
    for merged_range in merged_cells_ranges:
        sheet.merge_cells(str(merged_range))



    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

if st.button('Save Changes to Excel'):
    modified_excel = save_and_load_excel()
    st.download_button(
        label="Download Modified Excel File",
        data=modified_excel,
        file_name="Modified_Living_Wage_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )